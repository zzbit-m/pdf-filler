import atexit
import os
import shutil
import tempfile
from pathlib import Path

import fitz
import openpyxl
import pytest

_TEST_DATA_DIR = tempfile.mkdtemp(prefix="pbeam-test-")
os.environ.setdefault("PBEAM_DATA_DIR", _TEST_DATA_DIR)


@atexit.register
def _cleanup_test_data_dir() -> None:
    shutil.rmtree(_TEST_DATA_DIR, ignore_errors=True)


@pytest.fixture
def sample_pdf(tmp_path: Path) -> Path:
    path = tmp_path / "sample.pdf"
    doc = fitz.open()
    doc.new_page(width=595, height=842)
    page = doc[0]
    page.insert_text(fitz.Point(50, 100), "Test PDF", fontsize=12)
    doc.save(str(path))
    doc.close()
    return path


@pytest.fixture
def sample_excel(tmp_path: Path) -> Path:
    path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"
    ws.cell(row=3, column=1, value="Name")
    ws.cell(row=3, column=2, value="Position")
    ws.cell(row=3, column=3, value="Department")
    ws.cell(row=4, column=1, value="John Doe")
    ws.cell(row=4, column=2, value="Engineer")
    ws.cell(row=4, column=3, value="Engineering")
    ws.cell(row=5, column=1, value="Jane Smith")
    ws.cell(row=5, column=2, value="Manager")
    ws.cell(row=5, column=3, value="HR")
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture
def sample_font() -> str:
    return str(Path(__file__).resolve().parent.parent / "app" / "fonts" / "tahoma.ttf")


@pytest.fixture
def templates_dir(tmp_path: Path) -> Path:
    return tmp_path / "templates"

