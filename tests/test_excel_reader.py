from pathlib import Path

import pytest

from app.services import excel_reader


def test_read_columns(sample_excel: Path) -> None:
    columns = excel_reader.read_columns(sample_excel)
    assert columns == ["Name", "Position", "Department"]


def test_read_rows(sample_excel: Path) -> None:
    columns, rows = excel_reader.read_rows(sample_excel)
    assert columns == ["Name", "Position", "Department"]
    assert len(rows) == 2
    assert rows[0]["Name"] == "John Doe"
    assert rows[0]["Position"] == "Engineer"
    assert rows[0]["Department"] == "Engineering"
    assert rows[1]["Name"] == "Jane Smith"
    assert rows[1]["Position"] == "Manager"
    assert rows[1]["Department"] == "HR"


def test_skip_empty_rows(tmp_path: Path) -> None:
    path = tmp_path / "empty_rows.xlsx"
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(row=3, column=1, value="Col1")
    ws.cell(row=4, column=1, value="")
    ws.cell(row=5, column=1, value="")
    ws.cell(row=6, column=1, value="HasValue")
    wb.save(str(path))
    wb.close()

    columns, rows = excel_reader.read_rows(path)
    assert len(rows) == 1
    assert rows[0]["Col1"] == "HasValue"


def test_file_not_found() -> None:
    with pytest.raises(FileNotFoundError):
        excel_reader.read_columns("nonexistent.xlsx")


def test_empty_excel_raises(tmp_path: Path) -> None:
    import openpyxl
    path = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    wb.save(str(path))
    wb.close()
    columns = excel_reader.read_columns(path)
    assert columns == []


def test_read_unique_values(sample_excel: Path) -> None:
    values = excel_reader.read_unique_values(sample_excel, "Department")
    assert values == ["Engineering", "HR"]


def test_read_unique_values_column_not_found(sample_excel: Path) -> None:
    values = excel_reader.read_unique_values(sample_excel, "Nonexistent")
    assert values == []
