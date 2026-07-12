import io

import fitz
import openpyxl
from fastapi.testclient import TestClient

from app.config import DATA_BASE
from app.main import app
from app.routers.fill import fill_state

client = TestClient(app)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _create_excel(columns: list[str], data: list[list[str]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    for i, col in enumerate(columns, start=1):
        ws.cell(row=3, column=i, value=col)
    for r, row in enumerate(data, start=4):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _create_pdf(page_count: int = 1) -> bytes:
    doc = fitz.open()
    for _ in range(page_count):
        doc.new_page(width=595, height=842)
    data = doc.write()
    doc.close()
    return data


class TestExcelUpload:
    def test_upload_excel_success(self):
        content = _create_excel(["Name", "Role"], [["Alice", "Engineer"]])
        resp = client.post(
            "/upload/excel",
            files={"file": ("test.xlsx", content, XLSX_MIME)},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["columns"] == ["Name", "Role"]
        assert len(data["preview_rows"]) == 1
        assert data["preview_rows"][0]["Name"] == "Alice"

    def test_wrong_file_type(self):
        resp = client.post("/upload/excel", files={"file": ("test.txt", b"hello", "text/plain")})
        assert resp.status_code == 400
        assert "xlsx" in resp.json()["detail"].lower()

    def test_empty_columns(self):
        content = _create_excel([], [])
        resp = client.post(
            "/upload/excel",
            files={"file": ("empty.xlsx", content, XLSX_MIME)},
        )
        assert resp.status_code == 400


class TestPdfUpload:
    def test_upload_pdf_success(self):
        content = _create_pdf()
        resp = client.post("/upload/pdf", files={"file": ("a.pdf", content, "application/pdf")})
        assert resp.status_code == 200
        data = resp.json()
        assert "pdf_id" in data
        assert data["page_count"] == 1
        assert data["filename"] == "a.pdf"

    def test_wrong_file_type(self):
        resp = client.post("/upload/pdf", files={"file": ("test.txt", b"hello", "text/plain")})
        assert resp.status_code == 400
        assert "pdf" in resp.json()["detail"].lower()

    def test_corrupted_pdf(self):
        resp = client.post(
            "/upload/pdf",
            files={"file": ("bad.pdf", b"not a real pdf", "application/pdf")},
        )
        assert resp.status_code == 400
        assert "corrupted" in resp.json()["detail"].lower()

    def test_multi_page(self):
        content = _create_pdf(page_count=3)
        resp = client.post("/upload/pdf", files={"file": ("multi.pdf", content, "application/pdf")})
        assert resp.status_code == 200
        assert resp.json()["page_count"] == 3


class TestPreview:
    def test_preview_success(self):
        pdf_content = _create_pdf()
        upload_resp = client.post(
            "/upload/pdf",
            files={"file": ("a.pdf", pdf_content, "application/pdf")},
        )
        pdf_id = upload_resp.json()["pdf_id"]

        resp = client.get(f"/preview/{pdf_id}/1")
        assert resp.status_code == 200
        assert resp.headers["content-type"] == "image/png"

    def test_pdf_not_found(self):
        resp = client.get("/preview/00000000-0000-0000-0000-000000000000/1")
        assert resp.status_code == 404

    def test_page_out_of_range(self):
        pdf_content = _create_pdf()
        upload_resp = client.post(
            "/upload/pdf",
            files={"file": ("a.pdf", pdf_content, "application/pdf")},
        )
        pdf_id = upload_resp.json()["pdf_id"]

        resp = client.get(f"/preview/{pdf_id}/99")
        assert resp.status_code == 404


class TestTemplate:
    def test_save_and_list(self):
        fields = [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}]
        resp = client.post(
            "/template",
            json={"name": "Test", "pdf_file": "a.pdf", "fields": fields},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["name"] == "Test"
        assert data["field_count"] == 1
        template_id = data["id"]

        list_resp = client.get("/template/list")
        assert list_resp.status_code == 200
        ids = [t["id"] for t in list_resp.json()]
        assert template_id in ids

    def test_save_no_fields(self):
        resp = client.post(
            "/template",
            json={"name": "Empty", "pdf_file": "a.pdf", "fields": []},
        )
        assert resp.status_code == 400

    def test_get_and_delete(self):
        fields = [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}]
        create_resp = client.post(
            "/template",
            json={"name": "GetDel", "pdf_file": "a.pdf", "fields": fields},
        )
        tid = create_resp.json()["id"]

        get_resp = client.get(f"/template/{tid}")
        assert get_resp.status_code == 200
        assert get_resp.json()["name"] == "GetDel"

        del_resp = client.delete(f"/template/{tid}")
        assert del_resp.status_code == 200

        get_resp2 = client.get(f"/template/{tid}")
        assert get_resp2.status_code == 404

    def test_get_not_found(self):
        resp = client.get("/template/00000000-0000-0000-0000-000000000000")
        assert resp.status_code == 404

    def test_delete_not_found(self):
        resp = client.delete("/template/00000000-0000-0000-0000-000000000000")
        assert resp.status_code == 404

    def test_overlap_warning(self):
        fields = [
            {"column": "Name", "page": 1, "x": 100, "y": 100, "font_size": 11},
            {"column": "Role", "page": 1, "x": 101, "y": 101, "font_size": 11},
        ]
        resp = client.post(
            "/template",
            json={"name": "Overlap", "pdf_file": "a.pdf", "fields": fields},
        )
        assert resp.status_code == 200
        assert len(resp.json()["warnings"]) > 0

    def test_no_overlap_warning(self):
        fields = [
            {"column": "Name", "page": 1, "x": 100, "y": 100, "font_size": 11},
            {"column": "Role", "page": 1, "x": 300, "y": 300, "font_size": 11},
        ]
        resp = client.post(
            "/template",
            json={"name": "NoOverlap", "pdf_file": "a.pdf", "fields": fields},
        )
        assert resp.status_code == 200
        assert len(resp.json()["warnings"]) == 0

    def test_rename_template(self):
        fields = [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}]
        create = client.post(
            "/template",
            json={"name": "Old", "pdf_file": "a.pdf", "fields": fields},
        )
        tid = create.json()["id"]

        rename = client.put(f"/template/{tid}", json={"name": "Renamed"})
        assert rename.status_code == 200
        assert rename.json()["name"] == "Renamed"

        get = client.get(f"/template/{tid}")
        assert get.json()["name"] == "Renamed"

    def test_rename_template_404(self):
        resp = client.put("/template/00000000-0000-0000-0000-000000000000", json={"name": "X"})
        assert resp.status_code == 404

    def test_duplicate_template(self):
        fields = [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}]
        create = client.post(
            "/template",
            json={"name": "Original", "pdf_file": "a.pdf", "fields": fields},
        )
        tid = create.json()["id"]

        dup = client.post(f"/template/{tid}/duplicate", json={})
        assert dup.status_code == 200
        data = dup.json()
        assert data["id"] != tid
        assert data["name"] == "Original (Copy)"
        assert data["field_count"] == 1

    def test_duplicate_template_404(self):
        resp = client.post("/template/00000000-0000-0000-0000-000000000000/duplicate", json={})
        assert resp.status_code == 404

    def test_thumbnail(self):
        pdf_content = _create_pdf()
        pdf_resp = client.post(
            "/upload/pdf",
            files={"file": ("a.pdf", pdf_content, "application/pdf")},
        )
        pdf_id = pdf_resp.json()["pdf_id"]

        fields = [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}]
        tmpl_resp = client.post(
            "/template",
            json={"name": "Thumb", "pdf_file": f"{pdf_id}.pdf", "fields": fields},
        )
        tid = tmpl_resp.json()["id"]

        resp = client.get(f"/template/{tid}/thumbnail")
        assert resp.status_code == 200
        assert resp.headers["content-type"] == "image/png"

    def test_thumbnail_missing_pdf(self):
        fields = [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}]
        create = client.post(
            "/template",
            json={"name": "NoPdf", "pdf_file": "b.pdf", "fields": fields},
        )
        tid = create.json()["id"]
        resp = client.get(f"/template/{tid}/thumbnail")
        assert resp.status_code == 404

    def test_thumbnail_invalid_pdf_file(self):
        fields = [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}]
        resp = client.post(
            "/template",
            json={"name": "BadPdfRef", "pdf_file": "../../etc/passwd", "fields": fields},
        )
        assert resp.status_code == 400


class TestFill:
    def _upload_pdf_and_create_template(self, columns: list[str]) -> tuple[str, str]:
        pdf_content = _create_pdf()
        pdf_resp = client.post(
            "/upload/pdf",
            files={"file": ("template.pdf", pdf_content, "application/pdf")},
        )
        pdf_id = pdf_resp.json()["pdf_id"]

        fields = [
            {"column": col, "page": 1, "x": 100 + i * 50, "y": 200, "font_size": 11}
            for i, col in enumerate(columns)
        ]
        tmpl_resp = client.post(
            "/template",
            json={"name": "FillTest", "pdf_file": f"{pdf_id}.pdf", "fields": fields},
        )
        return tmpl_resp.json()["id"], pdf_id

    def test_full_flow(self):
        template_id, _ = self._upload_pdf_and_create_template(["Name"])
        excel_content = _create_excel(["Name"], [["Alice"], ["Bob"]])

        fill_resp = client.post(
            "/fill",
            params={"template_id": template_id},
            files={"file": ("data.xlsx", excel_content, XLSX_MIME)},
        )
        assert fill_resp.status_code == 200
        batch_id = fill_resp.json()["batch_id"]

        status_resp = client.get(f"/fill/{batch_id}/status")
        assert status_resp.status_code == 200
        status = status_resp.json()
        assert status["status"] == "completed"
        assert status["total"] == 2
        assert status["completed"] == 2

        dl_resp = client.get(f"/fill/{batch_id}/download")
        assert dl_resp.status_code == 200
        assert "zip" in dl_resp.headers["content-type"]

    def test_template_not_found(self):
        excel_content = _create_excel(["Name"], [["Alice"]])
        resp = client.post(
            "/fill",
            params={"template_id": "00000000-0000-0000-0000-000000000000"},
            files={"file": ("data.xlsx", excel_content, XLSX_MIME)},
        )
        assert resp.status_code == 404

    def test_wrong_file_type(self):
        template_id, _ = self._upload_pdf_and_create_template(["Name"])
        resp = client.post(
            "/fill",
            params={"template_id": template_id},
            files={"file": ("data.txt", b"hello", "text/plain")},
        )
        assert resp.status_code == 400

    def test_missing_columns_warning(self):
        template_id, _ = self._upload_pdf_and_create_template(["MissingCol"])
        excel_content = _create_excel(["Name"], [["Alice"]])

        resp = client.post(
            "/fill",
            params={"template_id": template_id},
            files={"file": ("data.xlsx", excel_content, XLSX_MIME)},
        )
        assert resp.status_code == 200
        assert len(resp.json()["warnings"]) > 0

    def test_status_not_found(self):
        resp = client.get("/fill/00000000-0000-0000-0000-000000000000/status")
        assert resp.status_code == 404

    def test_download_not_found(self):
        resp = client.get("/fill/00000000-0000-0000-0000-000000000000/download")
        assert resp.status_code == 404

    def test_download_not_completed(self):
        fill_state["pending-batch"] = {"status": "pending", "total": 5, "completed": 0}
        resp = client.get("/fill/pending-batch/download")
        assert resp.status_code == 400

    def test_fill_overlay_error(self):
        pdf_content = _create_pdf()
        pdf_resp = client.post(
            "/upload/pdf",
            files={"file": ("template.pdf", pdf_content, "application/pdf")},
        )
        pdf_id = pdf_resp.json()["pdf_id"]
        fields = [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}]
        tmpl_resp = client.post(
            "/template",
            json={"name": "ErrTest", "pdf_file": f"{pdf_id}.pdf", "fields": fields},
        )
        template_id = tmpl_resp.json()["id"]
        (DATA_BASE / "uploads" / f"{pdf_id}.pdf").unlink()
        excel_content = _create_excel(["Name"], [["Alice"]])
        fill_resp = client.post(
            "/fill",
            params={"template_id": template_id},
            files={"file": ("data.xlsx", excel_content, XLSX_MIME)},
        )
        assert fill_resp.status_code == 200
        batch_id = fill_resp.json()["batch_id"]
        status_resp = client.get(f"/fill/{batch_id}/status")
        assert status_resp.status_code == 200
        status = status_resp.json()
        assert status["status"] == "error"
        assert status["error"] is not None


class TestWorkflow:
    def test_create_workflow(self):
        fields = [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}]
        tmpl = client.post(
            "/template",
            json={"name": "WF Tmpl", "pdf_file": "a.pdf", "fields": fields},
        )
        tid = tmpl.json()["id"]

        resp = client.post("/workflow", json={
            "name": "Test WF",
            "routing_column": "Dept",
            "routes": [{"value": "Eng", "template_id": tid}],
        })
        assert resp.status_code == 200
        data = resp.json()
        assert data["name"] == "Test WF"
        assert data["routing_column"] == "Dept"
        assert data["route_count"] == 1

    def test_create_duplicate_route_values(self):
        fields = [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}]
        tmpl = client.post(
            "/template",
            json={"name": "WF Dup", "pdf_file": "a.pdf", "fields": fields},
        )
        tid = tmpl.json()["id"]

        resp = client.post("/workflow", json={
            "name": "Bad WF",
            "routing_column": "Dept",
            "routes": [
                {"value": "Eng", "template_id": tid},
                {"value": "Eng", "template_id": tid},
            ],
        })
        assert resp.status_code == 400

    def test_create_no_routes(self):
        resp = client.post("/workflow", json={
            "name": "Empty", "routing_column": "Dept", "routes": [],
        })
        assert resp.status_code == 422

    def test_list_workflows(self):
        fields = [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}]
        tmpl = client.post(
            "/template",
            json={"name": "WF List", "pdf_file": "a.pdf", "fields": fields},
        )
        tid = tmpl.json()["id"]
        client.post("/workflow", json={
            "name": "WF A", "routing_column": "C",
            "routes": [{"value": "x", "template_id": tid}],
        })

        list_resp = client.get("/workflow/list")
        assert list_resp.status_code == 200
        wfs = list_resp.json()
        assert len(wfs) >= 1
        names = [w["name"] for w in wfs]
        assert "WF A" in names

    def test_get_workflow(self):
        fields = [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}]
        tmpl = client.post(
            "/template",
            json={"name": "WF Get", "pdf_file": "a.pdf", "fields": fields},
        )
        tid = tmpl.json()["id"]
        create = client.post("/workflow", json={
            "name": "Get Me", "routing_column": "C",
            "routes": [{"value": "x", "template_id": tid}],
        })
        wid = create.json()["id"]

        get_resp = client.get(f"/workflow/{wid}")
        assert get_resp.status_code == 200
        data = get_resp.json()
        assert data["name"] == "Get Me"
        assert len(data["routes"]) == 1
        assert data["routes"][0]["template_name"] == "WF Get"

    def test_get_workflow_404(self):
        resp = client.get("/workflow/00000000-0000-0000-0000-000000000000")
        assert resp.status_code == 404

    def test_rename_workflow(self):
        fields = [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}]
        tmpl = client.post(
            "/template",
            json={"name": "WF Ren", "pdf_file": "a.pdf", "fields": fields},
        )
        tid = tmpl.json()["id"]
        create = client.post("/workflow", json={
            "name": "Old WF", "routing_column": "C",
            "routes": [{"value": "x", "template_id": tid}],
        })
        wid = create.json()["id"]

        rename = client.put(f"/workflow/{wid}", json={"name": "New WF"})
        assert rename.status_code == 200
        assert rename.json()["name"] == "New WF"

    def test_delete_workflow(self):
        fields = [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}]
        tmpl = client.post(
            "/template",
            json={"name": "WF Del", "pdf_file": "a.pdf", "fields": fields},
        )
        tid = tmpl.json()["id"]
        create = client.post("/workflow", json={
            "name": "Delete Me", "routing_column": "C",
            "routes": [{"value": "x", "template_id": tid}],
        })
        wid = create.json()["id"]

        del_resp = client.delete(f"/workflow/{wid}")
        assert del_resp.status_code == 200

        get_resp = client.get(f"/workflow/{wid}")
        assert get_resp.status_code == 404

    def test_workflow_fill_routing(self):
        pdf_A = _create_pdf()
        upload_A = client.post(
            "/upload/pdf",
            files={"file": ("a.pdf", pdf_A, "application/pdf")},
        )
        pdf_id_A = upload_A.json()["pdf_id"]

        pdf_B = _create_pdf()
        upload_B = client.post(
            "/upload/pdf",
            files={"file": ("b.pdf", pdf_B, "application/pdf")},
        )
        pdf_id_B = upload_B.json()["pdf_id"]

        tmpl_A = client.post("/template", json={
            "name": "Template A",
            "pdf_file": f"{pdf_id_A}.pdf",
            "fields": [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}],
        })
        tid_A = tmpl_A.json()["id"]

        tmpl_B = client.post("/template", json={
            "name": "Template B",
            "pdf_file": f"{pdf_id_B}.pdf",
            "fields": [{"column": "Role", "page": 1, "x": 100, "y": 200, "font_size": 11}],
        })
        tid_B = tmpl_B.json()["id"]

        wf_resp = client.post("/workflow", json={
            "name": "Routing WF",
            "routing_column": "Dept",
            "routes": [
                {"value": "Eng", "template_id": tid_A},
                {"value": "Sales", "template_id": tid_B},
            ],
        })
        wid = wf_resp.json()["id"]

        excel = _create_excel(
            ["Dept", "Name", "Role"],
            [["Eng", "Alice", "Engineer"], ["Sales", "Bob", "Sales Rep"]],
        )

        fill_resp = client.post(
            "/fill/workflow",
            params={"workflow_id": wid},
            files={"file": ("data.xlsx", excel, XLSX_MIME)},
        )
        assert fill_resp.status_code == 200
        batch_id = fill_resp.json()["batch_id"]

        status_resp = client.get(f"/fill/{batch_id}/status")
        assert status_resp.status_code == 200
        status = status_resp.json()
        assert status["status"] == "completed"
        assert status["total"] == 2
        assert status["completed"] == 2

        dl_resp = client.get(f"/fill/{batch_id}/download")
        assert dl_resp.status_code == 200
        assert "zip" in dl_resp.headers["content-type"]

    def test_workflow_fill_missing_routing_column(self):
        fields = [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}]
        tmpl = client.post(
            "/template",
            json={"name": "WF Col", "pdf_file": "a.pdf", "fields": fields},
        )
        tid = tmpl.json()["id"]
        wf = client.post("/workflow", json={
            "name": "Missing Col", "routing_column": "Dept",
            "routes": [{"value": "Eng", "template_id": tid}],
        })
        wid = wf.json()["id"]

        excel = _create_excel(["Name"], [["Alice"]])
        resp = client.post(
            "/fill/workflow",
            params={"workflow_id": wid},
            files={"file": ("data.xlsx", excel, XLSX_MIME)},
        )
        assert resp.status_code == 400
        assert "not found" in resp.json()["detail"].lower()

    def test_workflow_not_found(self):
        excel = _create_excel(["Name"], [["Alice"]])
        resp = client.post(
            "/fill/workflow",
            params={"workflow_id": "00000000-0000-0000-0000-000000000000"},
            files={"file": ("data.xlsx", excel, XLSX_MIME)},
        )
        assert resp.status_code == 404

    def test_workflow_fill_no_routes(self):
        fields = [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}]
        tmpl = client.post(
            "/template",
            json={"name": "WF NoR", "pdf_file": "a.pdf", "fields": fields},
        )
        tid = tmpl.json()["id"]
        create = client.post("/workflow", json={
            "name": "Empty Routes", "routing_column": "Dept",
            "routes": [{"value": "X", "template_id": tid}],
        })
        wid = create.json()["id"]
        del_resp = client.delete(f"/template/{tid}")
        assert del_resp.status_code == 200

        workflow_data = client.get(f"/workflow/{wid}").json()
        workflow_data["routes"] = []
        import json
        (DATA_BASE / "workflows" / f"{wid}.json").write_text(
            json.dumps(workflow_data, ensure_ascii=False, indent=2), encoding="utf-8",
        )

        excel = _create_excel(["Dept"], [["X"]])
        resp = client.post(
            "/fill/workflow",
            params={"workflow_id": wid},
            files={"file": ("data.xlsx", excel, XLSX_MIME)},
        )
        assert resp.status_code == 400

    def test_workflow_fill_deleted_template(self):
        pdf = _create_pdf()
        upload = client.post(
            "/upload/pdf",
            files={"file": ("t.pdf", pdf, "application/pdf")},
        )
        pdf_id = upload.json()["pdf_id"]
        tmpl = client.post("/template", json={
            "name": "Gone Soon", "pdf_file": f"{pdf_id}.pdf",
            "fields": [{"column": "Name", "page": 1, "x": 100, "y": 200, "font_size": 11}],
        })
        tid = tmpl.json()["id"]

        wf = client.post("/workflow", json={
            "name": "Dead Ref", "routing_column": "Dept",
            "routes": [{"value": "Eng", "template_id": tid}],
        })
        wid = wf.json()["id"]

        client.delete(f"/template/{tid}")

        excel = _create_excel(["Dept", "Name"], [["Eng", "Alice"]])
        fill_resp = client.post(
            "/fill/workflow",
            params={"workflow_id": wid},
            files={"file": ("data.xlsx", excel, XLSX_MIME)},
        )
        assert fill_resp.status_code == 200
        batch_id = fill_resp.json()["batch_id"]

        status_resp = client.get(f"/fill/{batch_id}/status")
        assert status_resp.status_code == 200
        status = status_resp.json()
        assert status["status"] == "completed"
        assert status["warnings"]
