from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

HEADER_ROW = 3


def _get_active_sheet(wb: openpyxl.Workbook) -> Worksheet:
    ws = wb.active
    assert ws is not None
    return ws


def _read_header_row(ws: Worksheet) -> list[Any]:
    try:
        return list(ws[HEADER_ROW])
    except IndexError:
        return []


def read_columns(filepath: str | Path) -> list[str]:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = _get_active_sheet(wb)
        columns = []
        for cell in _read_header_row(ws):
            if cell.value is not None:
                columns.append(str(cell.value).strip())
        return columns
    finally:
        wb.close()


def read_rows(filepath: str | Path) -> tuple[list[str], list[dict[str, str]]]:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = _get_active_sheet(wb)
        columns: list[str] = []
        col_positions: list[tuple[int, str]] = []
        for i, cell in enumerate(_read_header_row(ws)):
            if cell.value is not None:
                col_name = str(cell.value).strip()
                columns.append(col_name)
                col_positions.append((i, col_name))

        rows: list[dict[str, str]] = []
        for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            vals: tuple[Any, ...] = row
            row_dict: dict[str, str] = {}
            for col_idx, col_name in col_positions:
                if col_idx < len(vals) and vals[col_idx] is not None:
                    row_dict[col_name] = str(vals[col_idx]).strip()
                else:
                    row_dict[col_name] = ""
            if any(row_dict.values()):
                rows.append(row_dict)
        return columns, rows
    finally:
        wb.close()


def read_unique_values(filepath: str | Path, column_name: str) -> list[str]:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = _get_active_sheet(wb)
        col_idx: int | None = None
        for i, cell in enumerate(_read_header_row(ws)):
            if cell.value is not None and str(cell.value).strip() == column_name:
                col_idx = i
                break
        if col_idx is None:
            return []
        values: set[str] = set()
        for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            vals: tuple[Any, ...] = row
            if col_idx < len(vals) and vals[col_idx] is not None:
                v = str(vals[col_idx]).strip()
                if v:
                    values.add(v)
        return sorted(values)
    finally:
        wb.close()
